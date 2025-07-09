# apps/crew/views.py
from rest_framework import viewsets, status, filters
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from django.db.models import Q, Count, Prefetch
from django.utils import timezone
from django.shortcuts import get_object_or_404
import csv
import openpyxl
from io import StringIO
from datetime import datetime, timedelta

from .models import (
    Department, Position, CrewMember, CrewAssignment,
    CallSheet, CrewCall, Character
)
from .serializers import (
    DepartmentSerializer, PositionSerializer,
    CrewMemberListSerializer, CrewMemberDetailSerializer, CrewMemberCreateUpdateSerializer,
    CrewAssignmentListSerializer, CrewAssignmentDetailSerializer,
    CallSheetListSerializer, CallSheetDetailSerializer, CallSheetCreateSerializer,
    CrewCallSerializer, CharacterSerializer,
    CrewAvailabilitySerializer, CrewBulkImportSerializer
)

class DepartmentViewSet(viewsets.ModelViewSet):
    queryset = Department.objects.all()
    serializer_class = DepartmentSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [filters.OrderingFilter]
    ordering_fields = ['sort_order', 'name']
    ordering = ['sort_order']

class PositionViewSet(viewsets.ModelViewSet):
    queryset = Position.objects.select_related('department')
    serializer_class = PositionSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['title']
    ordering_fields = ['title', 'department__name']
    
    def get_queryset(self):
        queryset = super().get_queryset()
        department_id = self.request.query_params.get('department')
        if department_id:
            queryset = queryset.filter(department_id=department_id)
        return queryset

class CrewMemberViewSet(viewsets.ModelViewSet):
    queryset = CrewMember.objects.select_related('primary_position__department')
    permission_classes = [IsAuthenticated]
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['first_name', 'last_name', 'email', 'phone_primary']
    ordering_fields = ['last_name', 'first_name', 'created_at']
    ordering = ['last_name', 'first_name']
    
    def get_serializer_class(self):
        """Vrátí správný serializer podle akce"""
        if self.action in ['list']:
            return CrewMemberListSerializer
        elif self.action in ['create', 'update', 'partial_update']:
            return CrewMemberCreateUpdateSerializer
        else:
            return CrewMemberDetailSerializer
    
    def get_queryset(self):
        queryset = super().get_queryset()
        
        # Filter by department
        department_id = self.request.query_params.get('department')
        if department_id:
            queryset = queryset.filter(primary_position__department_id=department_id)
        
        # Filter by status
        status_param = self.request.query_params.get('status')
        if status_param:
            queryset = queryset.filter(status=status_param)
        
        # Filter by position
        position_id = self.request.query_params.get('position')
        if position_id:
            queryset = queryset.filter(primary_position_id=position_id)
        
        return queryset
    
    def create(self, request, *args, **kwargs):
        """Vytvoří nového crew member s detailním error handlingem"""
        serializer = self.get_serializer(data=request.data)
        
        if serializer.is_valid():
            try:
                crew_member = serializer.save()
                response_serializer = CrewMemberDetailSerializer(crew_member)
                return Response(
                    response_serializer.data, 
                    status=status.HTTP_201_CREATED
                )
            except Exception as e:
                return Response(
                    {'error': f'Failed to create crew member: {str(e)}'}, 
                    status=status.HTTP_400_BAD_REQUEST
                )
        else:
            return Response(
                {'errors': serializer.errors}, 
                status=status.HTTP_400_BAD_REQUEST
            )

    @action(detail=False, methods=['get'])
    def departments_with_positions(self, request):
        """Vrátí departments s jejich pozicemi pro dropdown"""
        departments = Department.objects.prefetch_related('positions').all()
        data = []
        
        for dept in departments:
            positions = [
                {
                    'id': str(pos.id),
                    'title': pos.title,
                    'level': pos.level,
                    'daily_rate_min': pos.daily_rate_min,
                    'daily_rate_max': pos.daily_rate_max
                }
                for pos in dept.positions.all()
            ]
            
            data.append({
                'id': dept.id,
                'name': dept.name,
                'abbreviation': dept.abbreviation,
                'color_code': dept.color_code,
                'positions': positions
            })
        
        return Response(data)
    
    @action(detail=False, methods=['post'])
    def check_availability(self, request):
        """Check crew availability for date range"""
        serializer = CrewAvailabilitySerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        
        data = serializer.validated_data
        start_date = data['start_date']
        end_date = data['end_date']
        
        # Get all crew members
        crew = CrewMember.objects.filter(status='active')
        
        # Filter by position/department if specified
        if 'position_id' in data:
            crew = crew.filter(
                Q(primary_position_id=data['position_id']) |
                Q(secondary_positions__id=data['position_id'])
            ).distinct()
        elif 'department_id' in data:
            crew = crew.filter(primary_position__department_id=data['department_id'])
        
        # Check assignments in date range
        available = []
        unavailable = []
        
        for member in crew:
            conflicts = member.assignments.filter(
                Q(start_date__lte=end_date) & Q(end_date__gte=start_date) |
                Q(start_date__lte=end_date) & Q(end_date__isnull=True),
                status__in=['confirmed', 'tentative']
            )
            
            if conflicts.exists():
                unavailable.append({
                    'crew_member': CrewMemberListSerializer(member).data,
                    'conflicts': CrewAssignmentListSerializer(conflicts, many=True).data
                })
            else:
                available.append(CrewMemberListSerializer(member).data)
        
        return Response({
            'available': available,
            'unavailable': unavailable,
            'total_available': len(available),
            'total_unavailable': len(unavailable)
        })
    
    @action(detail=False, methods=['post'])
    def bulk_import(self, request):
        """Import crew members from CSV/Excel file"""
        serializer = CrewBulkImportSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        
        file = serializer.validated_data['file']
        update_existing = serializer.validated_data['update_existing']
        
        imported = 0
        updated = 0
        errors = []
        
        try:
            if file.name.endswith('.csv'):
                # Process CSV
                decoded_file = file.read().decode('utf-8')
                reader = csv.DictReader(StringIO(decoded_file))
                
                for row in reader:
                    result = self._import_crew_member(row, update_existing)
                    if result['status'] == 'imported':
                        imported += 1
                    elif result['status'] == 'updated':
                        updated += 1
                    else:
                        errors.append(result['error'])
            
            else:
                # Process Excel
                wb = openpyxl.load_workbook(file)
                ws = wb.active
                headers = [cell.value for cell in ws[1]]
                
                for row in ws.iter_rows(min_row=2, values_only=True):
                    data = dict(zip(headers, row))
                    result = self._import_crew_member(data, update_existing)
                    if result['status'] == 'imported':
                        imported += 1
                    elif result['status'] == 'updated':
                        updated += 1
                    else:
                        errors.append(result['error'])
        
        except Exception as e:
            return Response(
                {'error': f'File processing error: {str(e)}'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        return Response({
            'imported': imported,
            'updated': updated,
            'errors': errors,
            'total_processed': imported + updated + len(errors)
        })
    
    def _import_crew_member(self, data, update_existing):
        """Import single crew member from data dict"""
        try:
            # Required fields
            email = data.get('email', '').strip()
            if not email:
                return {'status': 'error', 'error': 'Email is required'}
            
            # Check if exists
            existing = CrewMember.objects.filter(email=email).first()
            if existing and not update_existing:
                return {'status': 'error', 'error': f'{email} already exists'}
            
            # Prepare data
            crew_data = {
                'email': email,
                'first_name': data.get('first_name', '').strip(),
                'last_name': data.get('last_name', '').strip(),
                'phone_primary': data.get('phone', '').strip(),
                'emergency_contact_name': data.get('emergency_contact', '').strip(),
                'emergency_contact_phone': data.get('emergency_phone', '').strip(),
            }
            
            # Handle position
            position_title = data.get('position', '').strip()
            if position_title:
                position = Position.objects.filter(title__iexact=position_title).first()
                if position:
                    crew_data['primary_position'] = position
            
            if existing:
                for key, value in crew_data.items():
                    if value:  # Only update non-empty values
                        setattr(existing, key, value)
                existing.save()
                return {'status': 'updated'}
            else:
                CrewMember.objects.create(**crew_data)
                return {'status': 'imported'}
        
        except Exception as e:
            return {'status': 'error', 'error': str(e)}

class CrewAssignmentViewSet(viewsets.ModelViewSet):
    queryset = CrewAssignment.objects.select_related(
        'production', 'crew_member', 'position__department'
    )
    permission_classes = [IsAuthenticated]
    filter_backends = [filters.OrderingFilter]
    ordering_fields = ['start_date', 'created_at']
    ordering = ['-created_at']
    
    def get_serializer_class(self):
        if self.action == 'list':
            return CrewAssignmentListSerializer
        return CrewAssignmentDetailSerializer
    
    def get_queryset(self):
        queryset = super().get_queryset()
        
        # Filter by production
        production_id = self.request.query_params.get('production')
        if production_id:
            queryset = queryset.filter(production_id=production_id)
        
        # Filter by crew member
        crew_member_id = self.request.query_params.get('crew_member')
        if crew_member_id:
            queryset = queryset.filter(crew_member_id=crew_member_id)
        
        # Filter by status
        status_param = self.request.query_params.get('status')
        if status_param:
            queryset = queryset.filter(status=status_param)
        
        # Filter by date range
        date_from = self.request.query_params.get('date_from')
        date_to = self.request.query_params.get('date_to')
        if date_from:
            queryset = queryset.filter(start_date__gte=date_from)
        if date_to:
            queryset = queryset.filter(
                Q(end_date__lte=date_to) | Q(end_date__isnull=True)
            )
        
        return queryset
    
    @action(detail=False, methods=['post'])
    def bulk_assign(self, request):
        """Bulk assign crew to production"""
        production_id = request.data.get('production_id')
        assignments = request.data.get('assignments', [])
        
        if not production_id or not assignments:
            return Response(
                {'error': 'production_id and assignments required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        created = []
        errors = []
        
        for assignment_data in assignments:
            try:
                assignment = CrewAssignment.objects.create(
                    production_id=production_id,
                    **assignment_data
                )
                created.append(CrewAssignmentListSerializer(assignment).data)
            except Exception as e:
                errors.append({
                    'data': assignment_data,
                    'error': str(e)
                })
        
        return Response({
            'created': created,
            'errors': errors
        })

class CallSheetViewSet(viewsets.ModelViewSet):
    queryset = CallSheet.objects.select_related('production')
    permission_classes = [IsAuthenticated]
    filter_backends = [filters.OrderingFilter]
    ordering_fields = ['date', 'shooting_day']
    ordering = ['-date']
    
    def get_serializer_class(self):
        if self.action == 'list':
            return CallSheetListSerializer
        elif self.action == 'create':
            return CallSheetCreateSerializer
        return CallSheetDetailSerializer
    
    def get_queryset(self):
        queryset = super().get_queryset()
        
        # Filter by production
        production_id = self.request.query_params.get('production')
        if production_id:
            queryset = queryset.filter(production_id=production_id)
        
        # Filter by date range
        date_from = self.request.query_params.get('date_from')
        date_to = self.request.query_params.get('date_to')
        if date_from:
            queryset = queryset.filter(date__gte=date_from)
        if date_to:
            queryset = queryset.filter(date__lte=date_to)
        
        # Filter by status
        status_param = self.request.query_params.get('status')
        if status_param:
            queryset = queryset.filter(status=status_param)
        
        return queryset

class CharacterViewSet(viewsets.ModelViewSet):
    queryset = Character.objects.select_related('production', 'actor')
    serializer_class = CharacterSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['name']
    ordering_fields = ['name']
    ordering = ['name']
    
    def get_queryset(self):
        queryset = super().get_queryset()
        
        # Filter by production
        production_id = self.request.query_params.get('production')
        if production_id:
            queryset = queryset.filter(production_id=production_id)
        
        # Filter by type
        if self.request.query_params.get('principals'):
            queryset = queryset.filter(is_principal=True)
        if self.request.query_params.get('supporting'):
            queryset = queryset.filter(is_supporting=True)
        
        return queryset
